import openpyxl
import os
import glob

def search_xlsx(fpath):
    print(f"Searching {os.path.basename(fpath)}...")
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for col_idx, val in enumerate(row, start=1):
                    if val and any(term in str(val).lower() for term in ["disgrace", "national", "shame", "holes"]):
                        print(f"Match: Sheet: {sheet_name} | Cell: {openpyxl.utils.get_column_letter(col_idx)}{row_idx} -> {val}")
    except Exception as e:
        print(f"Error reading {fpath}: {e}")

def main():
    target_dir = r"c:\Users\HP\.gemini\antigravity-ide\scratch\osint-agent\tablet_documents"
    files = glob.glob(os.path.join(target_dir, "*.xlsx"))
    for fpath in files:
        search_xlsx(fpath)

if __name__ == "__main__":
    main()
